# -*- coding: utf-8 -*-
"""
Created on Wed Mar 11 21:37:42 2020

@author: yuanh
"""

from openpyxl import load_workbook
import numpy as np
import os
import csv

cwd = os.getcwd()

nVar = 10     # number of decision variables
nPop = 100      # population size 

Position = np.zeros((nPop, nVar))
Cost = np.zeros((nPop, 2))
wb = load_workbook('Positions.xlsx')
sheet = wb['1']

for i in range(0,100):
    for j in range(0,10):
        Position[i,j]=sheet.cell(row=i+2, column=j+2).value
        
for i in range(0,100):
    Cost[i,0] = Position[i,0]**2*1e-6
    Cost[i,1] = sheet.cell(row=i+2,column=12).value
        
        
Position_best = np.zeros((nPop, nVar))
Cost_best = np.zeros((nPop, 2))
wb = load_workbook('Positions.xlsx')
sheet = wb['0_PB']

for i in range(0,100):
    for j in range(0,10):
        Position_best[i,j]=sheet.cell(row=i+2, column=j+2).value
        
for i in range(0,100):
    Cost_best[i,0] = Position_best[i,0]**2*1e-6
    Cost_best[i,1] = sheet.cell(row=i+2,column=12).value
    

nRep = 27
Position_rep = np.zeros((nRep, nVar))
Cost_rep = np.zeros((nRep, 2))
wb = load_workbook('Positions.xlsx')
sheet = wb['0_REP']

for i in range(0,nRep):
    for j in range(0,10):
        Position_rep[i,j]=sheet.cell(row=i+2, column=j+2).value
        
for i in range(0,nRep):
    Cost_rep[i,0] = Position_rep[i,0]**2*1e-6
    Cost_rep[i,1] = sheet.cell(row=i+2,column=12).value
    
    
Velocity = np.zeros((nPop, nVar))
wb = load_workbook('Positions.xlsx')
sheet = wb['0_V']
for i in range(0,100):
    for j in range(0,10):
        Velocity[i,j]=sheet.cell(row=i+2, column=j+2).value
        
nMu = 100
index_mu = np.zeros((nMu,1))
wb = load_workbook('Positions.xlsx')
sheet = wb['0_mu']
for i in range(0,nMu):
    index_mu[i,0] = sheet.cell(row=i+2,column=1).value
    
    
nMu = 100
Position_mu = np.zeros((nMu, nVar))
Cost_mu = np.zeros((nMu, 2))
wb = load_workbook('Positions.xlsx')
sheet = wb['0_mu']

for i in range(0,nMu):
    for j in range(0,10):
        Position_mu[i,j]=sheet.cell(row=i+2, column=j+2).value
        
for i in range(0,nMu):
    Cost_mu[i,0] = Position_mu[i,0]**2*1e-6
    Cost_mu[i,1] = sheet.cell(row=i+2,column=12).value












it = -1
Cost = np.zeros((nPop, 2))
for i in range(nPop):
    Cost[i,0] = Position[i,0]**2*1e-6
    os.chdir(cwd+"\\"+str(it)+"_"+str(i))
    with open('case'+str(it)+'_'+str(i)+'_coupling_coefficient.csv') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        k = 0
        for row in csv_reader:
            if k == 1 and row[1]:
                Cost[i,1] = 1/float(row[1])
            k = k+1
os.chdir(cwd)