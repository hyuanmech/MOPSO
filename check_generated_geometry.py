# -*- coding: utf-8 -*-
"""
Created on Mon Mar 23 16:05:17 2020

@author: yuanh
"""

import os
import shutil
from openpyxl import load_workbook
import numpy as np

it = 6

flag = 0

nPop = 100

if flag == 1:
    n = 9
    index = np.zeros((n, 1))
    wb = load_workbook('Positions.xlsx')
    sheet = wb['2_mu']
    for i in range(n):
        index[i,0] = sheet.cell(row=i+2,column=1).value
    
if flag == 1:
    os.mkdir(str(it)+'_MU_all')
    for hh in range(n):
        source = "J:/Coupler_optimization/MOPSO_CAE/"+str(it)+"_"+str(int(index[hh,0]))+"_MU"+"/"+str(it)+"_"+str(int(index[hh,0]))+"_MU_geo.pdf"
        destination = "J:/Coupler_optimization/MOPSO_CAE/"+str(it)+"_MU_all"+"/"+str(it)+"_"+str(int(index[hh,0]))+"_MU_geo.pdf"
        shutil.copyfile(source, destination)
else:
    os.mkdir(str(it)+'_all')
    for hh in range(nPop):
        source = "J:/Coupler_optimization/MOPSO_CAE/"+str(it)+"_"+str(hh)+"/"+str(it)+"_"+str(hh)+"_geo.pdf"
        destination = "J:/Coupler_optimization/MOPSO_CAE/"+str(it)+"_all"+"/"+str(it)+"_"+str(hh)+"_geo.pdf"
        shutil.copyfile(source, destination)

    